from Directories import excel_sheet_directory
from openpyxl import load_workbook
from os import path
import numpy as np
import csv
import pandas as pd
from scipy import stats
from Setup.Maze import Maze
from PhysicsEngine.drawables import Arrow


def get_sheet():
    workbook = load_workbook(filename=excel_sheet_directory + path.sep + "Testable.xlsx")
    sheet = workbook.active
    return sheet


sheet = get_sheet()


class Forces:
    def __init__(self, humans, x):
        self.excel_index = humans.excel_index
        self.date = self.get_date()
        self.size = humans.size
        self.directory = self.force_directory()
        self.synchronization_offset(x.fps)
        self.filename = self.get_force_filename()
        self.occupied = humans.occupied
        self.abs_values = self.forces_loading(humans.frames, x.fps)
        self.angles = self.get_angles(humans, x)
        self.angles_load = self.angles - x.angle[:, np.newaxis]
        self.meters_load = self.get_meters_load(x)

    @staticmethod
    def get_meters_load(x):
        return Maze(x).force_attachment_positions()

    @staticmethod
    def get_angles(humans, x):
        from trajectory_inheritance.humans import angle_shift
        return humans.angles + \
               x.angle[:, np.newaxis] + \
               np.array(list(angle_shift[x.size].values()))[np.newaxis, :]

    def get_date(self):
        day = sheet.cell(row=self.excel_index, column=2).value
        daytime = sheet.cell(row=self.excel_index, column=6).value
        return day.replace(hour=daytime.hour, minute=daytime.minute)

    def force_directory(self):
        day_string = str(self.date.year) + '-' + str(self.date.month).zfill(2) + '-' + str(self.date.day).zfill(2)
        return ('{0}{1}phys-guru-cs{2}ants{3}Tabea{4}Human Experiments{5}Raw Data and Videos{7}'
                + day_string + '{8}Force Measurements{9}' + self.size).format(path.sep, path.sep, path.sep, path.sep,
                                                                              path.sep, path.sep, path.sep, path.sep,
                                                                              path.sep, path.sep)

    def synchronization_offset(self, fps):
        """ frame of turning on force meter relative to start of the raw movie """
        [minute, second] = [int(number) for number in
                            sheet.cell(row=self.excel_index, column=16).value.strip()[:-3].split(':')]
        frame_force_meter = (second + minute * 60) * fps

        """ if the frame of synchronization is BEFORE the start of the movie which was tracked """
        if sheet.cell(row=self.excel_index, column=16).value[0] == '-':
            frame_force_meter = - frame_force_meter

        """ time of tracking relative to start of the raw movie """
        frame_tracking = int(sheet.cell(row=self.excel_index, column=8).value)
        return frame_tracking - frame_force_meter

    def get_force_filename(self):
        txt_name = sheet.cell(row=self.excel_index, column=19).value
        if txt_name.endswith('.txt'):
            return txt_name
        elif txt_name == '/':
            return ''
        else:
            print('You still have to add the name of the force file in line ' + str(self.excel_index))
        return txt_name

    def forces_loading(self, frames, fps):
        from trajectory_inheritance.humans import participant_number
        # read force meter file
        with open(self.force_directory() + path.sep + self.filename, 'r') as f:
            reader = csv.reader(f, delimiter='\t')
            text_file_content = [line for line in reader]

        def convert_to_frames(fps, times):
            def correct_times():
                # for the large human SPT Maze, we have an additional two digits and an space in our txt file.
                # We have to get rid of this.
                for i, time in enumerate(times):
                    times[i] = [times[i][0].split(' ')[-1]]
                return times

            times = correct_times()
            seconds = [int(time[0].split(':')[0]) * 3600 + int(time[0].split(':')[1]) * 60 + int(time[0].split(':')[2])
                       for time
                       in times]
            seconds = [sec - seconds[0] for sec in seconds]
            frames = []
            for second in range(seconds[-1]):
                measurements_per_second = len([sec for sec in seconds if sec == second])
                for ii in range(measurements_per_second):
                    frames.append(second * fps + int(ii * fps / measurements_per_second))
            return frames

        sampled_frames = convert_to_frames(fps, text_file_content[1::2][1:-1])

        # load forces and set them relative to the baseline
        forces_txt = [[float(fu) for fu in fo[0].split(' ') if len(fu) > 1] for fo in text_file_content[0::2][:-1]]

        # all unoccupied force meters should have zero force
        empty_indices = [i for i in range(participant_number[self.size]) if i not in self.occupied]
        for empty_index in empty_indices:
            for j in range(len(forces_txt)):
                forces_txt[j][empty_index] = 0

        # every frame of the movie gets a force for every force meter
        forces_all_frames = []
        for frames_index in range(len(sampled_frames) - 1):
            for ii in range(sampled_frames[frames_index], sampled_frames[frames_index + 1]):
                forces_all_frames.append(forces_txt[frames_index])

        # find the offset of the first frame of the movie to the start of the force meter measurement
        synch_offset = self.synchronization_offset(fps)

        abs_values = []
        # write the force into the self.frame[:].forces variable
        for i, force_index in enumerate(range(synch_offset, len(frames) + synch_offset)):
            abs_values.append(forces_all_frames[force_index])
        return self.remove_force_outliers(np.array(abs_values))

    @staticmethod
    def remove_force_outliers(array):
        def remove_force_outliers_single_forcemeter(single):
            # only one measurement
            df_original = pd.DataFrame(single)

            outlier_index = np.where((np.abs(stats.zscore(df_original, axis=0)) < 5) == False)[0]
            df_original.values[outlier_index] = np.NaN
            df_no_outliers = df_original.interpolate()

            return np.array(df_no_outliers) - np.min(df_no_outliers)[0]

        return np.squeeze(np.apply_along_axis(remove_force_outliers_single_forcemeter, 0, array))

    def draw(self, display, x):
        force_attachments = display.my_maze.force_attachment_positions()
        for name in x.participants.occupied:
            self.arrow(display.i, force_attachments[name], name).draw(display)

    def arrow(self, i, force_meter_coor, name) -> Arrow:
        """
        :return: start, end and string for the display of the force as a triplet
        """
        start = force_meter_coor
        end = force_meter_coor + \
              self.abs_values[i, name] * np.array([np.cos(self.angles[i, name]), np.sin(self.angles[i, name])]) * 1 / 5
        return Arrow(np.array(start), np.array(end), str(name + 1))

    def part(self, name: int, reference_frame='maze') -> np.ndarray:
        """
        :param name: index of the participant
        :param reference_frame: 'maze' or 'load', dependent on desired reference frame
        :return: len(x.frames)x2 numpy.array with x and y components of the force vectors
        """
        if reference_frame == 'maze':
            a = self.angles[:, name]
        elif reference_frame == 'load':
            a = self.angles_load[:, name]
        else:
            raise ValueError('What frame of reference?')

        return np.array([np.cos(a), np.sin(a)]) * self.abs_values[:, name]

    # def debugger(human, forces_all_frames, x):
    #     if np.isnan(np.sum([human.frames[i].forces[1] for i in range(0, len(human.frames))])):
    #         with open(excel_sheet_directory + path.sep + 'mistakes.txt', 'a') as f:
    #             f.write('\n' + x.filename + '\n')
    #             f.write('original movie: ' + sheet.cell(row=human.excel_index, column=1).value + '\n')
    #             f.write('force file: ' + force_filename(human) + '\n')
    #             f.write(
    #                 'configuration time: ' + sheet.cell(row=human.excel_index, column=16).value + '\n')
    #             f.write('length of force measurement: ' +
    #                     str(int(np.floor(len(forces_all_frames) / x.fps / 60))).zfill(2) + ':' +
    #                     str(int(np.floor(len(forces_all_frames) / x.fps % 60))).zfill(2) + '\n')
    #             f.write('missing frames: ' +
    #                     str(len([i for i in range(len(human.frames))
    #                              if np.isnan(human.frames[i].forces[0])])) + '\n')
